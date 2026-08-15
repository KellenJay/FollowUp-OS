"""
generate_golden_dataset.py
HHH Golden Dataset Generator — AI Eval Builder skill

Generates a 3-sheet Excel golden dataset for any AI product following the
Helpful / Honest / Harmless evaluation framework.

Usage (JSON config file):
    python scripts/generate_golden_dataset.py --config my-config.json --output MyProduct_Golden_Dataset.xlsx

Usage (inline flags):
    python scripts/generate_golden_dataset.py \
        --product-name "MyProduct" \
        --product-description "One sentence description" \
        --app-model "google/gemini-2.5-pro" \
        --grader-model "gpt-4o-mini" \
        --dimensions '[...]' \
        --test-cases '[...]' \
        --output "MyProduct_Golden_Dataset.xlsx"

See references/config-schema.md for the full config format.
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


# ── Colour palette ────────────────────────────────────────────────
DARK_BG      = "0D1117"
MID_BG       = "161B22"
HEADER_BG    = "1F6FEB"
HELPFUL_BG   = "1A3A5C"
HONEST_BG    = "1A4A2E"
HARMLESS_BG  = "4A1A1A"
SUB_HELP     = "2D5F8A"
SUB_HON      = "2D7A4A"
SUB_HARM     = "7A2D2D"

FILLS = {
    "BAD":       ("3D1A1A", "F85149"),
    "AVERAGE":   ("3D3000", "E3B341"),
    "GOOD":      ("1A3D2A", "3FB950"),
    
}

WHITE       = "FFFFFF"
LIGHT_GREY  = "C9D1D9"
YELLOW      = "F0C000"


def _fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)


def _font(bold=False, sz=9, color=WHITE):
    return Font(name="Arial", bold=bold, size=sz, color=color)


def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border():
    s = Side(style="thin", color="30363D")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, ref, value, bg, font=None, align=None, border=True):
    c = ws[ref] if isinstance(ref, str) else ws.cell(*ref)
    c.value = value
    c.fill = _fill(bg)
    if font:
        c.font = font
    if align:
        c.alignment = align
    if border:
        c.border = _border()
    return c


# ── Sheet 1: Overview ─────────────────────────────────────────────
def build_overview(wb, cfg):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1F6FEB"
    ws.column_dimensions["A"].width = 28
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 20
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 24

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"{cfg['product_name']} — AI Eval Golden Dataset"
    c.font = Font(name="Arial", bold=True, size=20, color=WHITE)
    c.fill = _fill(HEADER_BG)
    c.alignment = _align()

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = (
        f"Eval framework: Helpful · Honest · Harmless (HHH)  |  "
        f"App model: {cfg.get('app_model','—')}  |  "
        f"Grader model: {cfg.get('grader_model','—')}  "
        f"(independent judge — different provider to avoid self-grading bias)"
    )
    c.font = Font(name="Arial", size=10, italic=True, color=LIGHT_GREY)
    c.fill = _fill(MID_BG)
    c.alignment = _align()

    sections = [
        ("", ""),
        ("WHAT IS THIS PRODUCT?", ""),
        ("Product", cfg["product_description"]),
        ("", ""),
        ("WHY A DIFFERENT MODEL FOR GRADING?", ""),
        ("Self-grading bias",
         f"The product runs on {cfg.get('app_model','the app model')}. Using the same model to grade "
         f"its own outputs produces scores clustering at 90–98 regardless of actual quality. "
         f"Switching the judge to {cfg.get('grader_model','an independent model')} (a different provider) "
         f"breaks the self-referential loop and produces realistic spread."),
        ("", ""),
        ("EVAL FRAMEWORK — HHH", ""),
        ("Helpful",  "Does the output solve the user's actual problem? Are key dimensions correct and detailed?"),
        ("Honest",   "Are claims truthful and well-attributed? Are sources specific and real — not vague labels?"),
        ("Harmless", "Does the output avoid presenting AI-synthesized content as verified fact? Is generated content clearly labelled?"),
        ("", ""),
        ("SCORING SCALE", ""),
        ("0–40",   "BAD — Output is empty, fabricated, or provides no usable signal for decisions."),
        ("41–70",  "AVERAGE — Useful output but notable weaknesses: vague sources, generic recommendations."),
        ("71–100", "GOOD — Production-ready output: specific sources, sensible priorities, actionable recommendations."),
    ]

    row = 3
    for label, content in sections:
        if label and not content:
            ws.merge_cells(f"A{row}:G{row}")
            c = ws[f"A{row}"]
            c.value = label
            c.font = Font(name="Arial", bold=True, size=11, color=YELLOW)
            c.fill = _fill(MID_BG)
            c.alignment = _align(h="left")
            ws.row_dimensions[row].height = 22
        elif label and content:
            ws[f"A{row}"].value = label
            ws[f"A{row}"].font = _font(bold=True)
            ws[f"A{row}"].fill = _fill(DARK_BG)
            ws[f"A{row}"].alignment = _align(h="left", v="top")
            ws[f"A{row}"].border = _border()
            ws.merge_cells(f"B{row}:G{row}")
            c = ws[f"B{row}"]
            c.value = content
            c.font = _font(color=LIGHT_GREY)
            c.fill = _fill(DARK_BG)
            c.alignment = _align(h="left", v="top")
            c.border = _border()
            ws.row_dimensions[row].height = 45
        else:
            ws.row_dimensions[row].height = 8
        row += 1


# ── Sheet 2: Golden Dataset ───────────────────────────────────────
def build_dataset(wb, cfg):
    ws = wb.create_sheet("Golden Dataset")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "3FB950"

    dimensions = cfg["dimensions"]  # list of {id, name, hhh, question, low, high}
    test_cases = cfg["test_cases"]  # list of {label, score, description, evals: {dim_id: text}}

    # Group dimensions by HHH bucket
    buckets = {"Helpful": [], "Honest": [], "Harmless": []}
    for d in dimensions:
        buckets[d["hhh"]].append(d)

    # Fixed columns: ID, Label, Input Type, Input Description, Ground Truth
    meta_cols = 5
    dim_cols = [d for bucket in buckets.values() for d in bucket]
    score_cols = 3  # Expected, Actual, Pass/Flag
    total_cols = meta_cols + len(dim_cols) + score_cols

    # Set widths
    widths = [6, 14, 16, 32, 32] + [22] * len(dim_cols) + [14, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1: Title
    ws.row_dimensions[1].height = 36
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value = f"{cfg['product_name']} — Golden Dataset (HHH Framework)"
    c.font = Font(name="Arial", bold=True, size=15, color=WHITE)
    c.fill = _fill(HEADER_BG)
    c.alignment = _align()

    # Row 2: HHH bucket headers
    ws.row_dimensions[2].height = 20
    col = meta_cols + 1
    for bucket, dims in buckets.items():
        if not dims:
            continue
        bg = {"Helpful": HELPFUL_BG, "Honest": HONEST_BG, "Harmless": HARMLESS_BG}[bucket]
        start = get_column_letter(col)
        end = get_column_letter(col + len(dims) - 1)
        ws.merge_cells(f"{start}2:{end}2")
        c = ws[f"{start}2"]
        c.value = bucket.upper()
        c.font = _font(bold=True, sz=10)
        c.fill = _fill(bg)
        c.alignment = _align()
        col += len(dims)

    # Score headers
    score_start = meta_cols + len(dim_cols) + 1
    ws.merge_cells(f"{get_column_letter(score_start)}2:{get_column_letter(total_cols)}2")
    c = ws[f"{get_column_letter(score_start)}2"]
    c.value = "SCORES"
    c.font = _font(bold=True, sz=10)
    c.fill = _fill(MID_BG)
    c.alignment = _align()

    # Row 3: Column sub-headers
    ws.row_dimensions[3].height = 60
    meta_headers = ["ID", "Quality\nLabel", "Input Type", "Input Description",
                    "Benchmark Answer\n(what correct output looks like)"]
    sub_bgs = {"Helpful": SUB_HELP, "Honest": SUB_HON, "Harmless": SUB_HARM}

    for i, h in enumerate(meta_headers, 1):
        c = ws.cell(row=3, column=i)
        c.value = h
        c.font = _font(bold=True, sz=8)
        c.fill = _fill(MID_BG)
        c.alignment = _align()
        c.border = _border()

    col = meta_cols + 1
    for d in dim_cols:
        bg = sub_bgs[d["hhh"]]
        c = ws.cell(row=3, column=col)
        c.value = f"{d['name']}\n{d['question']}"
        c.font = _font(bold=True, sz=7)
        c.fill = _fill(bg)
        c.alignment = _align()
        c.border = _border()
        col += 1

    for label, idx in [("Expected\nScore (0–100)", 0), ("Actual Score\n(fill in after grader runs)", 1),
                        ("Pass / Flag\n(within ±15 of expected?)", 2)]:
        c = ws.cell(row=3, column=score_start + idx)
        c.value = label
        c.font = _font(bold=True, sz=8)
        c.fill = _fill(MID_BG)
        c.alignment = _align()
        c.border = _border()

    # Data rows
    row = 4
    for tc in test_cases:
        lbl = tc["label"].upper()
        bg, lbl_color = FILLS.get(lbl, ("1C2128", WHITE))
        ws.row_dimensions[row].height = 85

        row_data = [
            f"TC-{row-3:02d}",
            lbl,
            tc.get("input_type", ""),
            tc.get("input_description", ""),
            tc.get("ground_truth", ""),
        ]
        for i, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=i, value=val)
            c.fill = _fill(bg)
            c.border = _border()
            if i == 2:
                c.font = Font(name="Arial", bold=True, size=10, color=lbl_color)
                c.alignment = _align()
            elif i == 1:
                c.font = _font(bold=True)
                c.alignment = _align()
            else:
                c.font = _font(color=LIGHT_GREY)
                c.alignment = _align(h="left", v="top")

        col = meta_cols + 1
        for d in dim_cols:
            c = ws.cell(row=row, column=col, value=tc.get("evals", {}).get(d["id"], ""))
            c.fill = _fill(bg)
            c.font = _font(sz=8, color=LIGHT_GREY)
            c.alignment = _align(h="left", v="top")
            c.border = _border()
            col += 1

        # Expected
        exp_col = score_start
        ws.cell(row=row, column=exp_col, value=tc["score"]).fill = _fill(bg)
        ws.cell(row=row, column=exp_col).font = _font(bold=True)
        ws.cell(row=row, column=exp_col).alignment = _align()
        ws.cell(row=row, column=exp_col).border = _border()

        # Actual (editable)
        act_col = score_start + 1
        ws.cell(row=row, column=act_col, value="").fill = _fill("1C2128")
        ws.cell(row=row, column=act_col).font = Font(name="Arial", bold=True, size=11, color=YELLOW)
        ws.cell(row=row, column=act_col).alignment = _align()
        ws.cell(row=row, column=act_col).border = _border()

        # Pass/Flag formula
        flag_col = score_start + 2
        exp_ref = f"{get_column_letter(exp_col)}{row}"
        act_ref = f"{get_column_letter(act_col)}{row}"
        ws.cell(row=row, column=flag_col,
                value=f'=IF({act_ref}="","—",IF(ABS({act_ref}-{exp_ref})<=15,"PASS","FLAG"))')
        ws.cell(row=row, column=flag_col).fill = _fill(bg)
        ws.cell(row=row, column=flag_col).font = _font(bold=True)
        ws.cell(row=row, column=flag_col).alignment = _align()
        ws.cell(row=row, column=flag_col).border = _border()

        row += 1

    # Legend
    ws.row_dimensions[row].height = 16
    ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
    c = ws[f"A{row}"]
    c.value = "PASS = actual score within ±15 of expected   |   FLAG = grader drift detected — review rubric or recalibrate"
    c.font = Font(name="Arial", italic=True, size=8, color=LIGHT_GREY)
    c.fill = _fill(MID_BG)
    c.alignment = _align()


# ── Sheet 3: Rubric ───────────────────────────────────────────────
def build_rubric(wb, cfg):
    ws = wb.create_sheet("Scoring Rubric")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "F0C000"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 38

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"{cfg['product_name']} — Scoring Rubric Reference"
    c.font = Font(name="Arial", bold=True, size=15, color=WHITE)
    c.fill = _fill(HEADER_BG)
    c.alignment = _align()

    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = "Use when assigning benchmark scores to new test cases, or auditing grader outputs."
    c.font = Font(name="Arial", italic=True, size=9, color=LIGHT_GREY)
    c.fill = _fill(MID_BG)
    c.alignment = _align()

    ws.row_dimensions[3].height = 28
    for i, h in enumerate(["Dimension", "LOW score (0–40)", "MEDIUM score (41–74)", "HIGH score (75–100)"], 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = _font(bold=True, sz=10)
        c.fill = _fill(MID_BG)
        c.alignment = _align()
        c.border = _border()

    bucket_bgs = {"Helpful": (HELPFUL_BG, SUB_HELP), "Honest": (HONEST_BG, SUB_HON), "Harmless": (HARMLESS_BG, SUB_HARM)}

    row = 4
    for d in cfg["dimensions"]:
        ws.row_dimensions[row].height = 70
        section_bg, row_bg = bucket_bgs[d["hhh"]]
        values = [
            d["name"],
            d.get("low", "Score 0–40: significant failure in this dimension."),
            d.get("medium", "Score 41–74: partial success with notable gaps."),
            d.get("high", "Score 75–100: strong performance on this dimension."),
        ]
        bgs = [section_bg, row_bg, row_bg, row_bg]
        for col, (val, bg) in enumerate(zip(values, bgs), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = _fill(bg)
            c.border = _border()
            if col == 1:
                c.font = _font(bold=True)
                c.alignment = _align()
            else:
                c.font = _font(sz=8, color=LIGHT_GREY)
                c.alignment = _align(h="left", v="top")
        row += 1


# ── Main ──────────────────────────────────────────────────────────
def load_config(args):
    if args.config:
        with open(args.config) as f:
            return json.load(f)

    # Build from individual flags
    cfg = {}
    cfg["product_name"] = args.product_name or "MyProduct"
    cfg["product_description"] = args.product_description or ""
    cfg["app_model"] = args.app_model or ""
    cfg["grader_model"] = args.grader_model or ""

    if args.dimensions:
        cfg["dimensions"] = json.loads(args.dimensions)
    else:
        cfg["dimensions"] = []

    if args.test_cases:
        cfg["test_cases"] = json.loads(args.test_cases)
    else:
        cfg["test_cases"] = []

    return cfg


def main():
    parser = argparse.ArgumentParser(description="Generate HHH Golden Dataset Excel file")
    parser.add_argument("--config", help="Path to JSON config file")
    parser.add_argument("--product-name")
    parser.add_argument("--product-description")
    parser.add_argument("--app-model")
    parser.add_argument("--grader-model")
    parser.add_argument("--dimensions", help="JSON array of dimension objects")
    parser.add_argument("--test-cases", help="JSON array of test case objects")
    parser.add_argument("--output", default="Golden_Dataset.xlsx")
    args = parser.parse_args()

    cfg = load_config(args)

    wb = openpyxl.Workbook()
    build_overview(wb, cfg)
    build_dataset(wb, cfg)
    build_rubric(wb, cfg)

    out = Path(args.output)
    wb.save(out)
    print(f"Saved: {out.resolve()}")


if __name__ == "__main__":
    main()
